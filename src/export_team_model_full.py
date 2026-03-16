from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def slugify(value: str) -> str:
    cleaned = str(value).strip()
    cleaned = cleaned.replace("%", " pct ")
    cleaned = cleaned.replace("/", " ")
    cleaned = cleaned.replace(",", " ")
    cleaned = cleaned.replace("(", " ").replace(")", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip().lower()
    cleaned = re.sub(r"[^a-z0-9]+", "_", cleaned)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned or "unnamed"


def build_headers(raw_headers: list) -> list[str]:
    headers = []
    seen = {}
    current_base = None
    sub_index = 0

    for idx, raw in enumerate(raw_headers, start=1):
        if raw is None or str(raw).strip() == "":
            if current_base is None:
                name = f"unnamed_{idx}"
            else:
                sub_index += 1
                name = f"{current_base}_sub{sub_index}"
        else:
            base = slugify(raw)
            sub_index = 0
            current_base = base
            count = seen.get(base, 0) + 1
            seen[base] = count
            name = base if count == 1 else f"{base}_{count}"
        headers.append(name)
    return headers


def main() -> None:
    input_path = Path("data/raw/USL C League Wide Performance & KPIs 2026 refined working.xlsx")
    output_path = Path("data/processed/team_model_full_export.csv")

    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb["Model"]

    raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = build_headers(raw_headers)

    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None or value == "" for value in row):
            continue
        data_rows.append(row)

    df = pd.DataFrame(data_rows, columns=headers)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(output_path, index=False)


if __name__ == "__main__":
    main()
